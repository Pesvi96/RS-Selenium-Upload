import openpyxl as xl
import os

"""
This module opens excel file and processes all values into a dictionary


My dictionary looks like this:
ID_LIST = {
    ID: [(price, purpose, vat),
        (price, purpose, vat),
        (price, purpose, vat)],
    ID: [(price, purpose, vat),
        (price, purpose, vat)]
    }
    
All values are as strings
"""

file_name = 'Logs/Working sheet - grouped data.xlsx'

current_path = os.getcwd()
file_path = current_path + "/" + file_name
workbook = xl.load_workbook(file_path)
worksheet = workbook.active


# Worksheet structure
PRICE_COLUMN = 4
ID_COLUMN = 1
PURPOSE_COLUMN = 3
VAT_COLUMN = 5
ROW_START = 2
row_range = 500

# List of all the values in the worksheet
ID_LIST = []
PRICE_LIST = []
PURPOSE_LIST = []
VAT_LIST = []


# Checks for all the values (Company ID, Price, Purpose, VAT (yes/no)) in the row.
# If all values are present (there is no blank data),
for row in range(ROW_START, row_range):
    id_value = worksheet.cell(row, ID_COLUMN).value
    price_value = worksheet.cell(row, PRICE_COLUMN).value
    purpose_value = worksheet.cell(row, PURPOSE_COLUMN).value
    vat_value = worksheet.cell(row, VAT_COLUMN).value
    # If all values are present in the worksheet row, append the values to corresponding list
    if id_value is not None and price_value is not None and purpose_value is not None and vat_value is not None:
        ID_LIST.append(str(id_value))
        PRICE_LIST.append(str(price_value))
        PURPOSE_LIST.append(str(purpose_value))
        VAT_LIST.append(str(vat_value))


MY_DICT = {}

# Append all values as a list of tuples under the corresponding (Company ID) key.
# Creates new key (Company ID) if it does not yet exist in the dictionary
for i in range(len(ID_LIST)):
    if ID_LIST[i] not in MY_DICT:
        MY_DICT.update({ID_LIST[i]: [(PRICE_LIST[i], PURPOSE_LIST[i], VAT_LIST[i])]})
    else:
        MY_DICT[ID_LIST[i]].append((PRICE_LIST[i], PURPOSE_LIST[i], VAT_LIST[i]))
        pass

workbook.close()



